from io import BytesIO
from pathlib import Path
import tempfile

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st

GOOGLE_SHEET_CSV = "https://docs.google.com/spreadsheets/d/1uDTjul2en8p0dLLpt964KyUjnIGW58LcZe_ZQBcmzww/export?format=csv&gid=0"

COMPETITOR_COLUMNS = [
    "All4Baby", "Amazon IE", "Baby Accessories", "Babymoon Baby Shop", "Bella Baby IE",
    "Boots IE", "Bundle Baby", "Byggebo", "Currys IE", "Ickle Bubba IE", "Kaliedy",
    "Kings baby shop", "Little Be", "Mamas And Papas IE", "mum NMe",
    "Next IE", "Rearfacing IE", "Samuel Johnston", "Smyths Toys IE", "Tony Kealys",
    "Tralee Nursery Supplies", "Very IE", "Wayfair IE"
]

ID_COLUMNS = [
    "Product name", "Baby Doc Shop", "Cost price", "SKU", "GTIN / EAN / UPC",
    "Category", "Brand", "My own price is..."
]

TEXT_COLUMNS = ["Product name", "SKU", "GTIN / EAN / UPC", "Category", "Brand", "My own price is..."]
NUMERIC_COLUMNS = ["Baby Doc Shop", "Cost price"] + COMPETITOR_COLUMNS

DISPLAY_COLUMNS = [
    "Product name",
    "Category",
    "BabyDoc Shop Price",
    "Babydocshop Margin %",
    "Brand",
    "Brand Price",
    "Competitor Margin %",
]

GREY_RGB_VALUES = {
    "FFBFBFBF", "00BFBFBF", "FFD9D9D9", "00D9D9D9", "FFCCCCCC", "00CCCCCC",
    "FF808080", "00808080", "FFA6A6A6", "00A6A6A6", "FF7F7F7F", "007F7F7F",
    "FFF2F2F2", "00F2F2F2", "FFE7E6E6", "00E7E6E6", "FFDBDBDB", "00DBDBDB",
    "FFDFDFDF", "00DFDFDF"
}

IGNORED_COMPETITOR_VALUES = {0, 0.0, "0", "0.0", "0.00", "€0", "€0.00"}


@st.cache_data(show_spinner=False)
def load_google_sheet():
    return pd.read_csv(GOOGLE_SHEET_CSV)


def is_grey_fill(cell):
    fill = cell.fill
    if fill is None:
        return False

    color_candidates = []
    for attr in ["fgColor", "start_color", "end_color"]:
        color_obj = getattr(fill, attr, None)
        if color_obj is not None:
            color_candidates.append(color_obj)

    for color_obj in color_candidates:
        rgb = getattr(color_obj, "rgb", None)
        if isinstance(rgb, str) and rgb.upper() in GREY_RGB_VALUES:
            return True

        indexed = getattr(color_obj, "indexed", None)
        if isinstance(indexed, int) and indexed in {15, 22, 23, 48, 55, 64}:
            return True

        theme = getattr(color_obj, "theme", None)
        tint = getattr(color_obj, "tint", None)
        color_type = getattr(color_obj, "type", None)

        try:
            if color_type == "theme" and theme is not None and tint is not None and float(tint) <= 0:
                return True
        except Exception:
            pass

    return False


def load_excel_without_grey_cells(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows())

    if not rows:
        return pd.DataFrame(), 0, 0

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in rows[0]]
    cleaned_rows = []
    grey_cells_removed = 0
    zero_cells_removed = 0
    competitor_header_set = {c.strip().lower() for c in COMPETITOR_COLUMNS}

    for row in rows[1:]:
        row_dict = {}
        for idx, cell in enumerate(row):
            header = headers[idx] if idx < len(headers) else f"Column_{idx + 1}"
            value = cell.value

            if header.strip().lower() in competitor_header_set:
                if is_grey_fill(cell):
                    value = None
                    grey_cells_removed += 1
                elif value in IGNORED_COMPETITOR_VALUES:
                    value = None
                    zero_cells_removed += 1
                else:
                    try:
                        numeric_value = float(value) if value is not None else None
                        if numeric_value == 0:
                            value = None
                            zero_cells_removed += 1
                    except Exception:
                        pass

            row_dict[header] = value
        cleaned_rows.append(row_dict)

    df = pd.DataFrame(cleaned_rows)
    df.columns = [str(c).strip() for c in df.columns]
    return df, grey_cells_removed, zero_cells_removed


def load_source(uploaded_file=None):
    grey_cells_removed = 0
    zero_cells_removed = 0

    if uploaded_file is not None:
        suffix = Path(uploaded_file.name).suffix.lower()
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name

        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            df, grey_cells_removed, zero_cells_removed = load_excel_without_grey_cells(tmp_path)
        else:
            df = pd.read_excel(tmp_path)
    else:
        df = load_google_sheet()

    df.columns = [str(c).strip() for c in df.columns]

    missing = [c for c in ID_COLUMNS + COMPETITOR_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    for col in NUMERIC_COLUMNS:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    for col in COMPETITOR_COLUMNS:
        df.loc[df[col].fillna(0) <= 0, col] = np.nan

    for col in TEXT_COLUMNS:
        df[col] = df[col].astype("string")

    return df, grey_cells_removed, zero_cells_removed


def get_vat_factor(category):
    cat = "" if pd.isna(category) else str(category).lower()
    if "car seat" in cat or "carseat" in cat or "car-seat" in cat or "isofix" in cat:
        return 1.13
    return 1.23


def calc_margin_percent(price_inc_vat, vat_factor, cost_price):
    if pd.isna(price_inc_vat) or pd.isna(cost_price) or price_inc_vat == 0:
        return np.nan
    price_ex_vat = price_inc_vat / vat_factor
    if price_ex_vat == 0:
        return np.nan
    return round(((price_ex_vat - cost_price) / price_ex_vat) * 100, 1)


def transform_data(df):
    melted = df.melt(
        id_vars=ID_COLUMNS,
        value_vars=COMPETITOR_COLUMNS,
        var_name="Competitor",
        value_name="Competitor Price",
    )

    melted["VAT Factor"] = melted["Category"].apply(get_vat_factor)
    melted["Value 1"] = np.where(
        melted["Competitor Price"].notna()
        & melted["Baby Doc Shop"].notna()
        & (melted["Competitor Price"] < melted["Baby Doc Shop"]),
        1,
        0,
    )

    melted["Babydocshop Margin %"] = melted.apply(
        lambda row: calc_margin_percent(row["Baby Doc Shop"], row["VAT Factor"], row["Cost price"]),
        axis=1,
    )
    melted["Competitor Margin %"] = melted.apply(
        lambda row: calc_margin_percent(row["Competitor Price"], row["VAT Factor"], row["Cost price"]),
        axis=1,
    )

    melted = melted[melted["Value 1"] == 1].copy()
    melted["Product Brand"] = melted["Brand"]
    melted["Brand"] = melted["Competitor"]
    melted["Brand Price"] = melted["Competitor Price"]
    melted["BabyDoc Shop Price"] = melted["Baby Doc Shop"]
    melted = melted.sort_values(["Product Brand", "Product name", "Competitor"], na_position="last")
    melted.reset_index(drop=True, inplace=True)
    return melted


def apply_filters(df, selected_product_brands, product_search, selected_competitors, selected_categories):
    out = df.copy()

    if selected_product_brands:
        out = out[out["Product Brand"].fillna("").isin(selected_product_brands)]

    if selected_competitors:
        out = out[out["Competitor"].fillna("").isin(selected_competitors)]

    if selected_categories:
        out = out[out["Category"].fillna("").isin(selected_categories)]

    if product_search and str(product_search).strip():
        search = str(product_search).strip().lower()
        out = out[out["Product name"].fillna("").str.lower().str.contains(search, na=False)]

    return out


def prepare_display(df):
    out = df[DISPLAY_COLUMNS].copy()
    for col in ["BabyDoc Shop Price", "Babydocshop Margin %", "Brand Price", "Competitor Margin %"]:
        out[col] = pd.to_numeric(out[col], errors="coerce").round(2)
    return out


def build_summary(filtered_display, filtered_full, grey_cells_removed=0, zero_cells_removed=0):
    grey_msg = f" | Grey cells removed: {grey_cells_removed:,}" if grey_cells_removed else ""
    zero_msg = f" | Zero-value cells removed: {zero_cells_removed:,}" if zero_cells_removed else ""
    return (
        f"Rows shown: {len(filtered_display):,} | "
        f"Products: {filtered_display['Product name'].nunique() if not filtered_display.empty else 0:,} | "
        f"Product brands: {filtered_full['Product Brand'].nunique() if not filtered_full.empty else 0:,} | "
        f"Competitor stores: {filtered_full['Competitor'].nunique() if not filtered_full.empty else 0:,}"
        f"{grey_msg}{zero_msg}"
    )


def dataframe_to_excel_bytes(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
    output.seek(0)
    return output


def main():
    st.set_page_config(page_title="Baby Doc Shop Price Monitor", layout="wide")

    st.title("Baby Doc Shop Price Monitor")
    st.write(
        "Upload your Matrix report from P2S or use the linked Google Sheet. "
        "This app keeps only rows where competitor product prices are lower than Baby Doc Shop, "
        "removes grey competitor cells and 0-value competitor cells, lets you select product names "
        "into a second table, and exports both tables to Excel."
    )

    uploaded_file = st.file_uploader("Upload daily Excel file", type=["xlsx", "xls", "xlsm", "xltx", "xltm"])
    use_google_sheet = st.checkbox("Use linked Google Sheet if no file is uploaded", value=True)

    if uploaded_file is None and not use_google_sheet:
        st.info("Upload an Excel file or tick the Google Sheet option.")
        return

    try:
        with st.spinner("Loading data..."):
            raw_df, grey_removed, zero_removed = load_source(uploaded_file)
            transformed = transform_data(raw_df)
    except Exception as exc:
        st.error(f"Error: {exc}")
        return

    st.success("Data loaded successfully.")

    product_brands = sorted([b for b in transformed["Product Brand"].dropna().astype(str).unique().tolist() if b.strip()])
    competitors = sorted([c for c in transformed["Competitor"].dropna().astype(str).unique().tolist() if c.strip()])
    categories = sorted([c for c in transformed["Category"].dropna().astype(str).unique().tolist() if c.strip()])

    with st.sidebar:
        st.header("Filters")
        selected_product_brands = st.multiselect("Product Brand", product_brands, default=product_brands)
        selected_competitors = st.multiselect("Competitor Store", competitors, default=competitors)
        selected_categories = st.multiselect("Category", categories, default=categories)
        product_search = st.text_input("Search product name")

    filtered_full = apply_filters(transformed, selected_product_brands, product_search, selected_competitors, selected_categories)
    preview = prepare_display(filtered_full)

    st.text(build_summary(preview, filtered_full, grey_removed, zero_removed))

    st.subheader("Main Table")
    st.dataframe(preview, use_container_width=True, hide_index=True)
    st.download_button(
        "Download main filtered table",
        data=dataframe_to_excel_bytes(preview),
        file_name="main_filtered_table.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=preview.empty,
    )

    st.subheader("Select Products For Change Request")
    product_names = sorted([p for p in filtered_full["Product name"].dropna().astype(str).unique().tolist() if p.strip()])
    selected_product_names = st.multiselect("Select product name(s)", product_names)

    if selected_product_names:
        selected_full = filtered_full[filtered_full["Product name"].isin(selected_product_names)].copy()
        selected_display = prepare_display(selected_full)
    else:
        selected_display = pd.DataFrame(columns=DISPLAY_COLUMNS)

    st.dataframe(selected_display, use_container_width=True, hide_index=True)
    st.download_button(
        "Download selected products table",
        data=dataframe_to_excel_bytes(selected_display),
        file_name="selected_products.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=selected_display.empty,
    )


if __name__ == "__main__":
    main()